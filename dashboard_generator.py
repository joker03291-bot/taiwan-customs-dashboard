"""
台湾関税統計 Dashboard ジェネレーター (v2: 汎用化版)

主要機能:
  - HS code・品名を原始資料から自動検出
  - 国名対照表を JSON で外部化、新規国家を自動補完
  - 品目別の単価妥当レンジを JSON 管理
  - 輸入/輸出いずれにも対応
  - すべての集計値は SUMIFS 等で総表から数式参照

使い方:
    from dashboard_generator import generate_dashboard

    result = generate_dashboard(
        input_path="raw_data.xls",
        output_path="dashboard.xlsx",
        trade_type="Imports",  # or "Exports", or None で自動判定
    )
    print(result["summary"])
"""

import json
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ 設定 ============

DATA_DIR = Path(__file__).parent / "data"
COUNTRY_MAPPING_FILE = DATA_DIR / "country_mapping.json"
COMMODITY_SETTINGS_FILE = DATA_DIR / "commodity_settings.json"

# スタイル定数
FONT_HEADER = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Yu Gothic", size=14, bold=True)
FONT_SUBTITLE = Font(name="Yu Gothic", size=10, italic=True, color="555555")
FONT_BODY = Font(name="Yu Gothic", size=10)
FONT_TOTAL = Font(name="Yu Gothic", size=10, bold=True)
FONT_NOTE = Font(name="Yu Gothic", size=9, color="555555")

FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_PROVISIONAL = PatternFill("solid", fgColor="FFF2CC")
FILL_RANK_TOP = PatternFill("solid", fgColor="E2EFDA")
FILL_NEW_COUNTRY = PatternFill("solid", fgColor="FCE4D6")  # 新規追加された国家用

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


# ============ 設定ファイル読み込み ============

def load_country_mapping() -> dict:
    """
    国名対照を読み込み。{en: {jp, order}} 形式で返す。
    alias (別名) も同じエントリへのキーとして登録される。
    例: 'Vietnam' (alias) と 'Viet Nam' (canonical) は両方とも 'ベトナム' を返す。
    """
    with open(COUNTRY_MAPPING_FILE, encoding="utf-8") as f:
        data = json.load(f)
    mapping = {}
    for entry in data["countries"]:
        if "en" in entry and "jp" in entry:
            info = {
                "jp": entry["jp"],
                "order": entry.get("order", 9999),
            }
            # canonical 名で登録
            mapping[entry["en"]] = info
            # alias も同じ info を指すように登録
            for alias in entry.get("aliases", []):
                if alias not in mapping:  # canonical 名と衝突しない場合のみ
                    mapping[alias] = info
    return mapping


def save_country_mapping(mapping: dict):
    """更新後の対照を JSON に保存 (新規国家追加時)。"""
    with open(COUNTRY_MAPPING_FILE, encoding="utf-8") as f:
        data = json.load(f)

    existing = {e["en"] for e in data["countries"] if "en" in e}
    for en, info in mapping.items():
        if en not in existing:
            data["countries"].append({
                "en": en,
                "jp": info["jp"],
                "order": info["order"],
                "auto_added": True,
                "auto_added_date": pd.Timestamp.now().strftime("%Y-%m-%d"),
            })

    data["_last_updated"] = pd.Timestamp.now().strftime("%Y-%m-%d")
    with open(COUNTRY_MAPPING_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_commodity_settings(commodity_code: str) -> dict:
    """品目コードに応じた設定を返す。未登録なら default。"""
    with open(COMMODITY_SETTINGS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    if commodity_code in data["commodities"]:
        return data["commodities"][commodity_code]
    return {
        "jp_name": None,  # 後で原始資料の英文名を使用
        **data["default"],
    }


# ============ 原始資料の解析 ============

def normalize_time(t):
    """
    時間文字列を正規化する。

    Returns:
        (clean: str, is_provisional: bool, is_monthly: bool)
        - clean: 括弧などを除去した文字列
        - is_provisional: 速報値かどうか
        - is_monthly: 月別形式 ('YYYY/M' or 'YYYY/MM') かどうか
    """
    s = str(t).strip()
    is_prelim = bool(re.search(r"preliminary|暫定|速報", s, re.IGNORECASE))
    clean = re.split(r"[(（]", s)[0].strip()
    # 月別形式の判定: YYYY/M または YYYY/MM のみを許可
    # '2024' (年のみ) や '2026/1~2026/3' (範囲) は不可
    is_monthly = bool(re.fullmatch(r"\d{4}/\d{1,2}", clean))
    return clean, is_prelim, is_monthly


def detect_commodity_info(df: pd.DataFrame) -> dict:
    """原始資料から品目コード・品名を抽出 (複数あればすべて返す)。"""
    codes = df["Commodity Code"].astype(str).unique()
    descs = df["Description"].astype(str).unique()
    return {
        "codes": list(codes),
        "descriptions": list(descs),
        "primary_code": str(codes[0]).strip(),
        "primary_desc": str(descs[0]).strip(),
        "is_multi": len(codes) > 1,
    }


def filter_trade_type(df: pd.DataFrame, trade_type: Optional[str]) -> tuple:
    """
    輸出入区分でフィルタ。trade_type=None なら自動判定 (両方あれば Imports 優先)。
    返り値: (filtered_df, actual_trade_type)
    """
    available = df["Imports/Exports"].unique()
    if trade_type is None:
        if "Imports" in available:
            trade_type = "Imports"
        elif "Exports" in available:
            trade_type = "Exports"
        else:
            trade_type = available[0]
    return df[df["Imports/Exports"] == trade_type].copy(), trade_type


def reconcile_countries(df: pd.DataFrame, mapping: dict) -> tuple:
    """
    原始資料の国家リストと対照表を照合。
    新規国家には自動で日文名を割り当て (英文名そのまま) +
    順序は既存最大値の次。返り値: (updated_mapping, new_countries_list)
    """
    raw_countries = df["Country"].dropna().unique()
    new_countries = []
    next_order = max(
        (info["order"] for info in mapping.values() if info["order"] < 1000),
        default=0,
    ) + 1

    for c in raw_countries:
        if c not in mapping:
            mapping[c] = {
                "jp": c,  # 暫定的に英文名のまま (ユーザーが後で編集)
                "order": next_order,
                "is_new": True,
            }
            new_countries.append(c)
            next_order += 1

    return mapping, new_countries


# ============ ワークブック構築 ============

def _setup_styles_for_cell(cell, font=FONT_BODY, fill=None, align=LEFT,
                           border=BORDER_ALL, num_fmt=None):
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = align
    cell.border = border
    if num_fmt:
        cell.number_format = num_fmt


def _sumifs_formula(value_col_letter: str, time_value: str,
                    country_jp: str, trade_type: str) -> str:
    """総表から SUMIFS で集計する数式を生成。"""
    return (f'=SUMIFS(総表!{value_col_letter}:{value_col_letter},'
            f'総表!A:A,"{trade_type}",'
            f'総表!B:B,"{time_value}",'
            f'総表!F:F,"{country_jp}")')


def _build_raw_sheet(wb, df, mapping):
    """ Sheet 1: 総表 (原始資料 + 日文国名)。 """
    ws = wb.create_sheet("総表")
    headers = ["輸出入区分", "年月", "品目コード", "品名", "国名(英)", "国名(日)",
               "金額(USD$1,000)", "重量(TNE)", "単価(USD/TNE)"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    df_sorted = df.reset_index(drop=True)
    for idx, row in df_sorted.iterrows():
        r = idx + 2
        en = row["Country"]
        jp = mapping.get(en, {}).get("jp", en)

        ws.cell(row=r, column=1, value=row["Imports/Exports"])
        ws.cell(row=r, column=2, value=row["Time"])
        ws.cell(row=r, column=3, value=str(row["Commodity Code"]))
        ws.cell(row=r, column=4, value=row["Description"])
        ws.cell(row=r, column=5, value=en)
        ws.cell(row=r, column=6, value=jp)
        ws.cell(row=r, column=7, value=float(row["Value(USD$1,000)"]))
        ws.cell(row=r, column=8, value=float(row["Weight(TNE)"]))
        ws.cell(row=r, column=9,
                value=f'=IF(H{r}=0,"",G{r}*1000/H{r})')

        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
            if col_idx in (7, 8):
                cell.number_format = "#,##0.00"
                cell.alignment = RIGHT
            elif col_idx == 9:
                cell.number_format = "#,##0.0"
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT if col_idx in (4, 5, 6) else CENTER

    ws.freeze_panes = "A2"
    widths = [12, 10, 14, 28, 24, 16, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_mapping_sheet(wb, mapping, used_countries, new_countries):
    """ Sheet 2: 国名対照 (使用中のもののみ)。 """
    ws = wb.create_sheet("国名対照")
    headers = ["順序", "国名(英)", "国名(日)", "ステータス", "備考"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    sorted_used = sorted(used_countries, key=lambda c: mapping[c]["order"])
    for idx, en in enumerate(sorted_used, start=1):
        info = mapping[en]
        is_new = en in new_countries
        status = "★ 新規" if is_new else "登録済"
        note = "原始資料に新規出現。日文名を編集してください。" if is_new else ""

        ws.cell(row=idx + 1, column=1, value=info["order"])
        ws.cell(row=idx + 1, column=2, value=en)
        ws.cell(row=idx + 1, column=3, value=info["jp"])
        ws.cell(row=idx + 1, column=4, value=status)
        ws.cell(row=idx + 1, column=5, value=note)

        for col_idx in range(1, 6):
            cell = ws.cell(row=idx + 1, column=col_idx)
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
            cell.alignment = CENTER if col_idx in (1, 4) else LEFT
            if is_new:
                cell.fill = FILL_NEW_COUNTRY

    ws.freeze_panes = "A2"
    for col_letter, w in zip("ABCDE", [8, 30, 18, 12, 35]):
        ws.column_dimensions[col_letter].width = w


def _build_dashboard_sheet(wb, used_countries, mapping, months,
                           provisional_months, trade_type, commodity_info,
                           commodity_settings, granularity="monthly"):
    """ ダッシュボードシート。granularity = 'monthly' or 'yearly'。 """
    trade_label = "輸入" if trade_type == "Imports" else "輸出"
    ws = wb.create_sheet(f"{trade_label}分析ダッシュボード")
    ws.sheet_view.showGridLines = False

    # 粒度に応じたラベル
    if granularity == "yearly":
        period_label = "年"           # 「月」→「年」
        period_sum_label = "年計"     # 「月計」→「年計」
        period_pivot_word = "年度別"  # 「月別」→「年度別」
        period_unit = "年"            # 「ヶ月」→「年」
        prev_period_label = "前年比"  # 「前月比」→「前年比」
    else:
        period_label = "月"
        period_sum_label = "月計"
        period_pivot_word = "月別"
        period_unit = "ヶ月"
        prev_period_label = "前月比"

    n_countries = len(used_countries)
    n_months = len(months)

    # 使用国 (jp) リスト, 順序順
    sorted_countries = sorted(used_countries, key=lambda c: mapping[c]["order"])
    jp_names = [mapping[c]["jp"] for c in sorted_countries]

    commodity_jp = commodity_settings.get("jp_name") or commodity_info["primary_desc"]
    commodity_code = commodity_info["primary_code"]
    price_min = commodity_settings["price_min"]
    price_max = commodity_settings["price_max"]

    prelim_str = ""
    if provisional_months:
        prelim_str = f" (※{'・'.join(sorted(provisional_months))}は速報値)"
    period_desc = f"{months[0]} 〜 {months[-1]}{prelim_str}"

    # --- タイトル ---
    ws.cell(row=1, column=2,
            value=f"{commodity_jp} {trade_label}分析ダッシュボード").font = FONT_TITLE
    ws.cell(row=2, column=2,
            value=(f"対象品目：{commodity_jp} (CCC: {commodity_code})　|　"
                   f"データ期間：{period_desc}　|　単位：金額 USD$1,000、重量 TNE　|　"
                   f"出所：台湾財政部関税統計")).font = FONT_SUBTITLE

    # --- 表 1: 期間 × 供給国 重量 ---
    t1_start = 4
    ws.cell(row=t1_start, column=2,
            value=f"■ 表1：{period_pivot_word}×{('供給国' if trade_type=='Imports' else '仕向国')} {trade_label}重量ピボット (TNE)").font = FONT_TOTAL

    t1_header = t1_start + 1
    label = "供給国" if trade_type == "Imports" else "仕向国"
    c = ws.cell(row=t1_header, column=2, value=label)
    _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    for i, m in enumerate(months):
        col = 3 + i
        is_p = str(m) in provisional_months
        c = ws.cell(row=t1_header, column=col, value=f"{m} *" if is_p else str(m))
        if is_p:
            _setup_styles_for_cell(c,
                                   font=Font(name="Yu Gothic", size=11, bold=True),
                                   fill=FILL_PROVISIONAL, align=CENTER)
        else:
            _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    total_col = 3 + n_months
    share_col = total_col + 1
    # ソート鍵欄: 視覚的に邪魔にならないよう、ダッシュボード本体から離れた位置 (列 30) に配置
    # 隠し列にすると Excel が分割線を表示するため、敢えて遠方に置いて見えにくくする
    sortkey_col = 30

    for col, txt in [(total_col, "合計"), (share_col, "構成比")]:
        c = ws.cell(row=t1_header, column=col, value=txt)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    t1_data_start = t1_header + 1
    for i, jp in enumerate(jp_names):
        r = t1_data_start + i
        c = ws.cell(row=r, column=2, value=jp)
        _setup_styles_for_cell(c, align=LEFT)

        for j, m in enumerate(months):
            col = 3 + j
            f = _sumifs_formula("H", m, jp, trade_type)
            c = ws.cell(row=r, column=col, value=f)
            _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        sl = get_column_letter(3); el = get_column_letter(3 + n_months - 1)
        c = ws.cell(row=r, column=total_col, value=f"=SUM({sl}{r}:{el}{r})")
        _setup_styles_for_cell(c, font=FONT_TOTAL, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        total_row_ref = t1_data_start + n_countries
        tl = get_column_letter(total_col)
        c = ws.cell(row=r, column=share_col,
                    value=f"=IFERROR({tl}{r}/${tl}${total_row_ref},0)")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0%")

    t1_total_row = t1_data_start + n_countries
    c = ws.cell(row=t1_total_row, column=2, value=period_sum_label)
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL, align=LEFT)

    for j, m in enumerate(months):
        col = 3 + j
        cl = get_column_letter(col)
        sr = t1_data_start; er = t1_data_start + n_countries - 1
        c = ws.cell(row=t1_total_row, column=col, value=f"=SUM({cl}{sr}:{cl}{er})")
        _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                               align=RIGHT, num_fmt="#,##0;-#,##0;-")

    tl = get_column_letter(total_col)
    er = t1_data_start + n_countries - 1
    c = ws.cell(row=t1_total_row, column=total_col,
                value=f"=SUM({tl}{t1_data_start}:{tl}{er})")
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                           align=RIGHT, num_fmt="#,##0;-#,##0;-")
    c = ws.cell(row=t1_total_row, column=share_col,
                value=f"={tl}{t1_total_row}/${tl}${t1_total_row}")
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                           align=RIGHT, num_fmt="0.0%")

    # --- 表 2: 月別 × 供給国 金額 (+ ソート鍵) ---
    t2_start = t1_total_row + 2
    ws.cell(row=t2_start, column=2,
            value=f"■ 表2：{period_pivot_word}×{label} {trade_label}金額ピボット (USD$1,000)").font = FONT_TOTAL

    t2_header = t2_start + 1
    c = ws.cell(row=t2_header, column=2, value=label)
    _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    for i, m in enumerate(months):
        col = 3 + i
        is_p = str(m) in provisional_months
        c = ws.cell(row=t2_header, column=col, value=f"{m} *" if is_p else str(m))
        if is_p:
            _setup_styles_for_cell(c,
                                   font=Font(name="Yu Gothic", size=11, bold=True),
                                   fill=FILL_PROVISIONAL, align=CENTER)
        else:
            _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    for col, txt in [(total_col, "合計"), (share_col, "構成比"),
                     (sortkey_col, "ソート鍵")]:
        c = ws.cell(row=t2_header, column=col, value=txt)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    t2_data_start = t2_header + 1
    for i, jp in enumerate(jp_names):
        r = t2_data_start + i
        c = ws.cell(row=r, column=2, value=jp)
        _setup_styles_for_cell(c, align=LEFT)

        for j, m in enumerate(months):
            col = 3 + j
            f = _sumifs_formula("G", m, jp, trade_type)
            c = ws.cell(row=r, column=col, value=f)
            _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        sl = get_column_letter(3); el = get_column_letter(3 + n_months - 1)
        c = ws.cell(row=r, column=total_col, value=f"=SUM({sl}{r}:{el}{r})")
        _setup_styles_for_cell(c, font=FONT_TOTAL, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        total_row_ref = t2_data_start + n_countries
        tl = get_column_letter(total_col)
        c = ws.cell(row=r, column=share_col,
                    value=f"=IFERROR({tl}{r}/${tl}${total_row_ref},0)")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0%")

        # ソート鍵
        offset = (n_countries - i) * 1e-5
        c = ws.cell(row=r, column=sortkey_col,
                    value=f"={tl}{r}+{offset}")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.00000")

    t2_total_row = t2_data_start + n_countries
    c = ws.cell(row=t2_total_row, column=2, value=period_sum_label)
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL, align=LEFT)

    for j, m in enumerate(months):
        col = 3 + j
        cl = get_column_letter(col)
        sr = t2_data_start; er = t2_data_start + n_countries - 1
        c = ws.cell(row=t2_total_row, column=col, value=f"=SUM({cl}{sr}:{cl}{er})")
        _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                               align=RIGHT, num_fmt="#,##0;-#,##0;-")

    tl = get_column_letter(total_col)
    er = t2_data_start + n_countries - 1
    c = ws.cell(row=t2_total_row, column=total_col,
                value=f"=SUM({tl}{t2_data_start}:{tl}{er})")
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                           align=RIGHT, num_fmt="#,##0;-#,##0;-")
    c = ws.cell(row=t2_total_row, column=share_col,
                value=f"={tl}{t2_total_row}/${tl}${t2_total_row}")
    _setup_styles_for_cell(c, font=FONT_TOTAL, fill=FILL_TOTAL,
                           align=RIGHT, num_fmt="0.0%")

    sortkey_col_letter = get_column_letter(sortkey_col)
    # ソート鍵列はダッシュボード本体から離れているため、隠す必要なし
    # (隠すと Excel が分割線を表示してしまうため)

    # --- 表 3: ランキング ---
    t3_start = t2_total_row + 3
    ws.cell(row=t3_start, column=2,
            value=f"■ 表3：{label}ランキングと構成比").font = FONT_TOTAL

    t3_header = t3_start + 1
    headers_t3 = ["順位", label, "総額(USD$1k)", "総重量(TNE)",
                  "金額構成比", "累計構成比", "平均単価(USD/TNE)", "備考"]
    for col_idx, h in enumerate(headers_t3, start=2):
        c = ws.cell(row=t3_header, column=col_idx, value=h)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    total_col_letter = get_column_letter(total_col)
    t2_amounts_range = f"{total_col_letter}{t2_data_start}:{total_col_letter}{t2_data_start + n_countries - 1}"
    t2_names_range = f"B{t2_data_start}:B{t2_data_start + n_countries - 1}"
    t1_weights_range = f"{total_col_letter}{t1_data_start}:{total_col_letter}{t1_data_start + n_countries - 1}"
    grand_total_amount = f"{total_col_letter}{t2_total_row}"
    sortkey_range = f"{sortkey_col_letter}{t2_data_start}:{sortkey_col_letter}{t2_data_start + n_countries - 1}"

    t3_data_start = t3_header + 1
    for i in range(n_countries):
        r = t3_data_start + i
        rank = i + 1

        c = ws.cell(row=r, column=2, value=rank)
        _setup_styles_for_cell(c, align=CENTER)

        sortkey_formula = f"LARGE({sortkey_range},{rank})"
        c = ws.cell(row=r, column=3,
                    value=f'=IFERROR(INDEX({t2_names_range},MATCH({sortkey_formula},{sortkey_range},0)),"")')
        _setup_styles_for_cell(c, align=LEFT)

        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(INDEX({t2_amounts_range},MATCH({sortkey_formula},{sortkey_range},0)),0)')
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        c = ws.cell(row=r, column=5,
                    value=f'=IFERROR(INDEX({t1_weights_range},MATCH({sortkey_formula},{sortkey_range},0)),0)')
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        c = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/{grand_total_amount},0)")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0%")

        if i == 0:
            c = ws.cell(row=r, column=7, value=f"=F{r}")
        else:
            c = ws.cell(row=r, column=7, value=f"=G{r-1}+F{r}")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0%")

        c = ws.cell(row=r, column=8, value=f"=IFERROR(D{r}*1000/E{r},0)")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0;-0.0;-")

        # 備考
        if price_min == 0 and price_max >= 999999:
            note_formula = f'=IF(E{r}=0,"※微量・要確認","")'
        else:
            note_formula = (f'=IF(E{r}=0,"※微量・要確認",'
                            f'IF(OR(H{r}<{price_min},H{r}>{price_max}),"※単価要確認",""))')
        c = ws.cell(row=r, column=9, value=note_formula)
        _setup_styles_for_cell(c, font=FONT_NOTE, align=LEFT)

        if rank <= 3:
            for col_idx in range(2, 10):
                ws.cell(row=r, column=col_idx).fill = FILL_RANK_TOP

    t3_last_row = t3_data_start + n_countries - 1

    # 上位 3 サマリー
    t3_summary_row = t3_last_row + 1
    ws.cell(row=t3_summary_row, column=2,
            value=(f'=" ▶ 上位3カ国（"&C{t3_data_start}&"・"&C{t3_data_start+1}&"・"'
                   f'&C{t3_data_start+2}&"）で全体の "&TEXT(G{t3_data_start+2},"0.0%")'
                   f'&" を占める。"')).font = FONT_NOTE

    # --- 表 4: 期間別単価推移 ---
    t4_start = t3_summary_row + 2
    ws.cell(row=t4_start, column=2,
            value=f"■ 表4：{period_pivot_word}平均単価推移 (USD/TNE) と{prev_period_label}").font = FONT_TOTAL

    t4_header = t4_start + 1
    headers_t4 = [period_label, f"{trade_label}金額(USD$1k)",
                  f"{trade_label}重量(TNE)", "平均単価(USD/TNE)", prev_period_label, "備考"]
    for col_idx, h in enumerate(headers_t4, start=2):
        c = ws.cell(row=t4_header, column=col_idx, value=h)
        _setup_styles_for_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align=CENTER)

    t4_data_start = t4_header + 1
    sheet_name = f"{trade_label}分析ダッシュボード"
    for i, m in enumerate(months):
        r = t4_data_start + i
        is_p = str(m) in provisional_months
        c = ws.cell(row=r, column=2, value=f"{m} *" if is_p else str(m))
        _setup_styles_for_cell(c, align=CENTER, fill=FILL_PROVISIONAL if is_p else None)

        month_col_letter = get_column_letter(3 + i)
        c = ws.cell(row=r, column=3,
                    value=f"='{sheet_name}'!{month_col_letter}{t2_total_row}")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        c = ws.cell(row=r, column=4,
                    value=f"='{sheet_name}'!{month_col_letter}{t1_total_row}")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="#,##0;-#,##0;-")

        c = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}*1000/D{r},0)")
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="0.0;-0.0;-")

        if i == 0:
            c = ws.cell(row=r, column=6, value="")
        else:
            c = ws.cell(row=r, column=6, value=f'=IFERROR((E{r}-E{r-1})/E{r-1},"")')
        _setup_styles_for_cell(c, align=RIGHT, num_fmt="+0.0%;-0.0%;-")

        c = ws.cell(row=r, column=7, value="速報値" if is_p else "")
        _setup_styles_for_cell(c, font=FONT_NOTE, align=LEFT)

    t4_last_row = t4_data_start + n_months - 1

    # --- 注記 ---
    note_start = t4_last_row + 2
    ws.cell(row=note_start, column=2, value="【注記】").font = FONT_TOTAL
    notes = [
        f"・ 出所：台湾財政部関税統計（CCC: {commodity_code}）",
        f"・ 区分：{trade_label} ({trade_type})  |  単位：金額 USD$1,000、重量 TNE",
        "・ 「総表」シートが全データの原本。表1〜4の数値はすべて SUMIFS 等の数式で総表から自動集計。",
        "・ 総表のデータを差し替えれば、ダッシュボード全体が自動再計算されます。",
        "・ 「–」表示は微量端数（< 1 TNE / < 1 USD$1k）の四捨五入により発生したゼロ。",
        "・ 「*」付きの月は速報値。確定値発表後に総表を更新してください。",
    ]
    if not (price_min == 0 and price_max >= 999999):
        notes.append(f"・ 表3「備考」欄の「※単価要確認」は平均単価が市況レンジ（{price_min}〜{price_max} USD/TNE）外。")
    notes.append("・ 国名対照は「国名対照」シートで管理。")
    if commodity_info["is_multi"]:
        notes.append(f"・ ※注意: 原始資料に複数の品目コード ({len(commodity_info['codes'])} 件) が含まれています。本ダッシュボードはすべてを合算しています。")

    for i, note in enumerate(notes):
        ws.cell(row=note_start + 1 + i, column=2, value=note).font = FONT_NOTE

    # 列幅
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    for col_idx in range(3, total_col + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    ws.freeze_panes = "C6"


# ============ メインAPI ============

def generate_dashboard(input_path: str, output_path: str,
                       trade_type=None,
                       update_country_mapping_file: bool = False) -> dict:
    """
    関税統計の原始ファイルからダッシュボードを生成。

    Args:
        input_path: 関税統計の .xls ファイルパス
        output_path: 出力 .xlsx ファイルパス
        trade_type: None (自動判定) | "Imports" | "Exports" |
                    ["Imports", "Exports"] (両方を別シートに出力)
        update_country_mapping_file: 新規国家を JSON 対照に永続化するか

    Returns:
        dict: 処理結果のサマリー
    """
    # --- 読み込み (xls / xlsx 自動判別) ---
    ext = Path(input_path).suffix.lower()
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    df_raw = pd.read_excel(input_path, engine=engine, header=0)
    df_raw.columns = ["Imports/Exports", "Time", "Commodity Code", "Description",
                      "Country", "Value(USD$1,000)", "Weight(TNE)"]

    # 時間正規化 (3-tuple: clean, is_provisional, is_monthly)
    parsed = df_raw["Time"].map(normalize_time)
    df_raw["Time_clean"] = [p[0] for p in parsed]
    df_raw["Is_provisional"] = [p[1] for p in parsed]
    df_raw["Is_monthly"] = [p[2] for p in parsed]
    df_raw["Time"] = df_raw["Time_clean"]
    df_raw = df_raw.drop(columns=["Time_clean"])

    # 粒度を判定 (monthly / yearly / mixed)
    monthly_count = df_raw["Is_monthly"].sum()
    total_count = len(df_raw)
    if monthly_count == total_count:
        granularity = "monthly"
    elif monthly_count == 0:
        granularity = "yearly"
    else:
        # 混合 — 通常起こらないが防御的に
        raise ValueError(
            f"原始ファイルに月別と非月別の時間値が混在しています。"
            f"月別: {monthly_count}件、非月別: {total_count - monthly_count}件。"
            f"財政部関税統計でダウンロードする際、月別または年別のいずれか一方を選択してください。"
        )

    # 「合計」「Total」などのダミー国家行を除去 (年度資料に多い)
    dummy_country_patterns = re.compile(
        r"^(total|grand[\s_-]*total|合計|総計|小計|計|all\s*countries?)$",
        re.IGNORECASE,
    )
    before = len(df_raw)
    df_raw = df_raw[~df_raw["Country"].astype(str).str.strip().str.match(dummy_country_patterns, na=False)].copy()
    removed_dummy_rows = before - len(df_raw)

    provisional_months = set(df_raw[df_raw["Is_provisional"]]["Time"].unique())

    # 区分の解決: trade_type を list[str] に正規化
    available = list(df_raw["Imports/Exports"].unique())
    if trade_type is None:
        # 自動判定: 利用可能なものすべてを使う
        trade_types = [t for t in ["Imports", "Exports"] if t in available]
        if not trade_types:
            trade_types = available[:1]
    elif isinstance(trade_type, str):
        trade_types = [trade_type]
    else:
        trade_types = list(trade_type)

    # データに存在しない区分を除去
    trade_types = [t for t in trade_types if t in available]
    if not trade_types:
        raise ValueError(
            f"指定された区分 {trade_type} は原始データに存在しません。"
            f"利用可能: {available}"
        )

    # 各区分の DataFrame をまとめる (総表は両方を含む)
    df_by_trade = {t: df_raw[df_raw["Imports/Exports"] == t].copy() for t in trade_types}
    df_combined = pd.concat(df_by_trade.values(), ignore_index=True)

    # 品目情報 (両区分の品目を統合)
    commodity_info = detect_commodity_info(df_combined)
    commodity_settings = load_commodity_settings(commodity_info["primary_code"])

    # 期間リスト (粒度に応じて並び替え) — 両区分を統合
    def _period_sort_key(s):
        """月別/年度の両形式に対応した並べ替えキー。"""
        s = str(s)
        m = re.fullmatch(r"(\d{4})/(\d{1,2})", s)
        if m:
            return (int(m.group(1)), int(m.group(2)), 0)
        m = re.fullmatch(r"(\d{4})/(\d{1,2})~\d{4}/\d{1,2}", s)
        if m:
            return (int(m.group(1)), int(m.group(2)), 1)
        m = re.fullmatch(r"(\d{4})", s)
        if m:
            return (int(m.group(1)), 0, 2)
        return (9999, 99, 9)

    # 国名対照 — 両区分の国家を統合
    mapping = load_country_mapping()
    mapping, new_countries = reconcile_countries(df_combined, mapping)
    used_countries_all = list(df_combined["Country"].dropna().unique())

    # --- ワークブック構築 ---
    wb = Workbook()
    wb.remove(wb.active)

    # 総表: 両区分すべて
    _build_raw_sheet(wb, df_combined, mapping)
    _build_mapping_sheet(wb, mapping, used_countries_all, new_countries)

    # ダッシュボードを区分ごとに構築
    per_trade_summaries = []
    for t in trade_types:
        df_t = df_by_trade[t]
        months_t = sorted(df_t["Time"].unique(), key=_period_sort_key)
        used_t = list(df_t["Country"].dropna().unique())
        _build_dashboard_sheet(wb, used_t, mapping, months_t,
                               provisional_months, t, commodity_info,
                               commodity_settings, granularity)
        per_trade_summaries.append({
            "trade_type": t,
            "n_periods": len(months_t),
            "n_countries": len(used_t),
            "n_records": len(df_t),
        })

    # シート順序: ダッシュボード → 総表 → 国名対照
    # 後ろから前に move するため, 逆順で処理
    for t in reversed(trade_types):
        trade_label = "輸入" if t == "Imports" else "輸出"
        wb.move_sheet(f"{trade_label}分析ダッシュボード", offset=-len(wb.sheetnames) + 1)

    wb.save(output_path)

    if update_country_mapping_file and new_countries:
        save_country_mapping(mapping)

    # サマリー文字列の組立
    summary_parts = []
    for s in per_trade_summaries:
        label = "輸入" if s["trade_type"] == "Imports" else "輸出"
        summary_parts.append(
            f"{label} {s['n_periods']}{'年' if granularity == 'yearly' else 'ヶ月'}×{s['n_countries']}カ国 ({s['n_records']}件)"
        )
    summary = (f"{commodity_settings.get('jp_name') or commodity_info['primary_desc']} "
               f"({commodity_info['primary_code']}) | " + " / ".join(summary_parts))

    # 統合期間 (表示用)
    months_all = sorted(df_combined["Time"].unique(), key=_period_sort_key)

    return {
        "status": "success",
        "output_path": output_path,
        "trade_types": trade_types,
        "trade_type": trade_types[0] if len(trade_types) == 1 else "Both",
        "per_trade_summaries": per_trade_summaries,
        "commodity_code": commodity_info["primary_code"],
        "commodity_jp": commodity_settings.get("jp_name") or commodity_info["prim
