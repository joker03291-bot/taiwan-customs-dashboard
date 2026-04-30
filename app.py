"""
台湾関税統計 Dashboard ジェネレーター (Streamlit Web UI)

起動: streamlit run app.py
"""

import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st

from dashboard_generator import (
    generate_dashboard,
    load_country_mapping,
    load_commodity_settings,
    normalize_time,
    detect_commodity_info,
)
from _theme import inject_theme, hero, footer

# ========== ページ設定 ==========

st.set_page_config(
    page_title="関税統計ダッシュボード生成ツール",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

inject_theme()

# ========== ヘッダー (Hero) ==========

hero(
    eyebrow="TAIWAN CUSTOMS STATISTICS",
    title="関税統計ダッシュボード自動生成ツール",
    desc=(
        "台湾財政部関税署の輸出入統計データから、"
        "日本語版の分析ダッシュボードを自動で生成します。"
        "毎月の供給先報告にそのままご利用いただけます。"
    ),
    cta="↓ 下記の手順に沿って操作してください",
)

# ========== サイドバー ==========

with st.sidebar:
    st.markdown("### マスタ管理")
    st.markdown(
        """
        左メニューから以下のページに移動できます：

        - **国名対照** — 英文国名と日文国名の対照表
        - **品目設定** — HS code ごとの日文品名・単価レンジ
        """
    )
    st.divider()
    st.caption("v2.0 / 2026-04")

# ========== 使用ステップ ==========

st.markdown("### 操作手順")

st.markdown(
    """
    <div class="timeline">
      <div class="step">
        <div class="step-number">1</div>
        <div class="step-title">原始データの取得</div>
        <div class="step-body">
          台湾財政部関税署の統計サイトから .xls ファイルをダウンロードします。<br>
          <a href="https://portal.sw.nat.gov.tw/APGA/GA03" target="_blank" rel="noopener">
            財政部関税署 統計ポータル を開く ↗
          </a>
        </div>
        <div class="step-note">
          <strong>取得方法：</strong>
          「綜合查詢」を選択 → HS code を入力 → 期間を指定 → .xls 形式でダウンロード
          <table>
            <thead>
              <tr><th>形式</th><th>内容</th><th>推奨</th></tr>
            </thead>
            <tbody>
              <tr><td><strong>月別</strong></td><td>各月ごとの輸出入数量・金額</td><td>◎ 通常はこちら</td></tr>
              <tr><td><strong>年度別</strong></td><td>年単位の集計値（月別ピボット不可）</td><td>長期トレンド分析時</td></tr>
            </tbody>
          </table>
        </div>
      </div>

      <div class="step">
        <div class="step-number">2</div>
        <div class="step-title">ファイルのアップロード</div>
        <div class="step-body">
          下のエリアに .xls ファイルをドラッグ＆ドロップしてください。
          自動で品目・期間・国家リストを検出します。
        </div>
      </div>

      <div class="step">
        <div class="step-number">3</div>
        <div class="step-title">輸出入区分の選択</div>
        <div class="step-body">
          輸入（Imports）／輸出（Exports）のいずれか、または両方を選択します。
        </div>
      </div>

      <div class="step">
        <div class="step-number">4</div>
        <div class="step-title">ダッシュボードのダウンロード</div>
        <div class="step-body">
          生成された .xlsx をダウンロードし、そのまま供給先へ送付できます。
        </div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.divider()

# ========== ファイルアップロード ==========

st.markdown("### ファイルアップロード")
st.caption("関税署からダウンロードした .xls ファイルを選択してください。")

uploaded = st.file_uploader(
    label="ファイル選択",
    type=["xls", "xlsx"],
    label_visibility="collapsed",
)

if uploaded is None:
    st.markdown(
        """
        <div class="info-card">
            <div class="info-card-title">ファイルがアップロードされていません</div>
            <p class="info-card-body">
                上のエリアに関税署からダウンロードした .xls ファイルをドラッグするか、
                クリックしてファイルを選択してください。
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("原始データの形式について"):
        st.markdown(
            """
            関税署のダウンロード形式（long-format）をそのまま使用します：

            | Imports/Exports | Time | Commodity Code | Description | Country | Value(USD$1,000) | Weight(TNE) |
            |---|---|---|---|---|---|---|
            | Imports | 2025/1 | 27101964009 | Naphtha, mineral | Russian Federation | 119,197 | 199,892 |
            | Imports | 2025/1 | 27101964009 | Naphtha, mineral | Singapore | 13,453 | 22,178 |

            速報値の月は `2026/3(preliminary)` のように `(preliminary)` が付与されます。
            """
        )

    footer()
    st.stop()

# ========== ファイル解析 ==========

ext = Path(uploaded.name).suffix.lower()
with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
    tmp.write(uploaded.read())
    tmp_path = tmp.name

# 原始データ読み込み
try:
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    df_raw = pd.read_excel(tmp_path, engine=engine, header=0)
    df_raw.columns = ["Imports/Exports", "Time", "Commodity Code", "Description",
                      "Country", "Value(USD$1,000)", "Weight(TNE)"]
except Exception as e:
    st.error(
        "ファイルの読み込みに失敗しました。\n\n"
        "**確認事項：**\n"
        "- 関税署からダウンロードしたファイルですか？\n"
        "- ファイル形式は .xls または .xlsx ですか？\n"
        "- ファイルが破損していませんか？（一度開いて確認してみてください）\n\n"
        "問題が続く場合は、新しいファイルをダウンロードし直してお試しください。"
    )
    with st.expander("技術的な詳細（管理者向け）"):
        st.code(str(e))
    footer()
    st.stop()

# 列が空の場合
if len(df_raw) == 0:
    st.warning(
        "ファイルにデータが含まれていません。\n\n"
        "関税署のサイトで条件を変更し、再度ダウンロードしてください。"
    )
    footer()
    st.stop()

# 時間正規化
parsed = df_raw["Time"].map(normalize_time)
df_raw["Time_clean"] = [p[0] for p in parsed]
df_raw["Is_provisional"] = [p[1] for p in parsed]
df_raw["Is_monthly"] = [p[2] for p in parsed]
df_raw["Time"] = df_raw["Time_clean"]

# 粒度判定
monthly_count = df_raw["Is_monthly"].sum()
total_count = len(df_raw)
if monthly_count == total_count:
    detected_granularity = "monthly"
    granularity_label = "月別"
elif monthly_count == 0:
    detected_granularity = "yearly"
    granularity_label = "年度別"
else:
    st.error(
        "**時間値の形式が混在しています。**\n\n"
        f"月別形式：{monthly_count} 件　／　非月別形式：{total_count - monthly_count} 件\n\n"
        "本ツールは「月別のみ」または「年度別のみ」のいずれか単一形式に対応しています。\n\n"
        "**対処方法：**\n"
        "関税署サイトでダウンロードする際、「月別」または「年度別」のいずれか一方のみを"
        "選択してダウンロードし直してください。"
    )
    footer()
    st.stop()

# ========== 検出結果 (簡潔表示) ==========

st.markdown("### 検出結果")

# 並べ替え用キー
import re as _re_local
def _sk(s):
    s = str(s)
    m = _re_local.fullmatch(r"(\d{4})/(\d{1,2})", s)
    if m: return (int(m.group(1)), int(m.group(2)), 0)
    m = _re_local.fullmatch(r"(\d{4})/(\d{1,2})~\d{4}/\d{1,2}", s)
    if m: return (int(m.group(1)), int(m.group(2)), 1)
    m = _re_local.fullmatch(r"(\d{4})", s)
    if m: return (int(m.group(1)), 0, 2)
    return (9999, 99, 9)

months = sorted(df_raw["Time"].unique(), key=_sk)
n_records = len(df_raw)
n_countries = df_raw["Country"].nunique()
trade_types_in_data = df_raw["Imports/Exports"].unique()
period_str = f"{months[0]} 〜 {months[-1]}" if len(months) > 0 else "-"

# 主要情報を一行サマリーで提示
st.markdown(
    f"""
    <div class="section-card">
      <div class="section-label">FILE SUMMARY</div>
      <div style="display: flex; gap: 2.5rem; flex-wrap: wrap; margin-top: 0.4rem;">
        <div>
          <div style="color:#64748b; font-size:0.78rem;">対象期間</div>
          <div style="color:#1a202c; font-size:1.05rem; font-weight:600;">{period_str}</div>
        </div>
        <div>
          <div style="color:#64748b; font-size:0.78rem;">粒度</div>
          <div style="color:#1a202c; font-size:1.05rem; font-weight:600;">{granularity_label}</div>
        </div>
        <div>
          <div style="color:#64748b; font-size:0.78rem;">レコード数</div>
          <div style="color:#1a202c; font-size:1.05rem; font-weight:600;">{n_records:,}</div>
        </div>
        <div>
          <div style="color:#64748b; font-size:0.78rem;">国家数</div>
          <div style="color:#1a202c; font-size:1.05rem; font-weight:600;">{n_countries}</div>
        </div>
        <div>
          <div style="color:#64748b; font-size:0.78rem;">区分</div>
          <div style="color:#1a202c; font-size:1.05rem; font-weight:600;">{' / '.join(trade_types_in_data)}</div>
        </div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# 詳細情報は折りたたみ
with st.expander("詳細情報を表示", expanded=False):
    # 粒度の通知
    if detected_granularity == "yearly":
        st.info(
            "**年度別データを検出しました。** "
            "各表の時間軸は「年」単位で集計されます（月別ピボット・前月比は表示されません）。"
        )

    # 品目情報
    df_for_commodity = df_raw[df_raw["Imports/Exports"].isin(trade_types_in_data)]
    commodity_info = detect_commodity_info(df_for_commodity)
    commodity_code = commodity_info["primary_code"]
    commodity_settings = load_commodity_settings(commodity_code)

    st.markdown("**品目情報**")
    col_p1, col_p2 = st.columns(2)
    with col_p1:
        st.text(f"HS code: {commodity_code}")
        st.text(f"英文品名: {commodity_info['primary_desc']}")
    with col_p2:
        jp_name = commodity_settings.get("jp_name") or commodity_info["primary_desc"]
        is_registered = commodity_settings.get("jp_name") is not None
        st.text(f"日文品名: {jp_name}")
        if is_registered:
            st.text(
                f"単価レンジ: {commodity_settings['price_min']}〜"
                f"{commodity_settings['price_max']} USD/TNE"
            )
        else:
            st.warning(
                "この品目は未登録です。「品目設定」ページで登録すると、"
                "日文名と単価市況レンジが反映されます。"
            )

    if commodity_info["is_multi"]:
        st.warning(
            f"原始データに複数の品目コード ({len(commodity_info['codes'])} 件) が"
            "含まれています。本ツールはすべて合算して処理します。"
        )

    # 国家対照チェック
    mapping = load_country_mapping()
    raw_countries_in_data = set(df_for_commodity["Country"].dropna().unique())
    unmapped = raw_countries_in_data - set(mapping.keys())
    if unmapped:
        st.markdown("**新規国家の検出**")
        st.warning(
            f"対照表に未登録の国家が **{len(unmapped)} 件** あります： "
            f"{', '.join(sorted(unmapped))}\n\n"
            "生成は可能ですが、これらの国家は英文名のまま表示されます。"
            "日文名を設定するには「国名対照」ページで編集してください。"
        )

# ========== 設定セクション ==========

st.markdown("### 生成設定")

col_a, col_b = st.columns(2)

with col_a:
    available_types = list(df_raw["Imports/Exports"].unique())
    has_imports = "Imports" in available_types
    has_exports = "Exports" in available_types

    st.markdown("**輸出入区分** （複数選択可）")
    cb_col1, cb_col2 = st.columns(2)
    with cb_col1:
        sel_imports = st.checkbox(
            "Imports（輸入）",
            value=has_imports,
            disabled=not has_imports,
            help="原始データに含まれていない場合は選択不可",
        )
    with cb_col2:
        sel_exports = st.checkbox(
            "Exports（輸出）",
            value=has_exports and not has_imports,
            disabled=not has_exports,
        )

    selected_trades = []
    if sel_imports:
        selected_trades.append("Imports")
    if sel_exports:
        selected_trades.append("Exports")

    if not selected_trades:
        st.warning("少なくとも 1 つの区分を選択してください。")

with col_b:
    if selected_trades:
        df_filtered = df_raw[df_raw["Imports/Exports"].isin(selected_trades)]
        period_display = (
            f"{df_filtered['Time'].min()} 〜 {df_filtered['Time'].max()}"
        )
    else:
        df_filtered = df_raw
        period_display = ""
    st.text_input("対象期間", value=period_display, disabled=True)

# ========== 生成ボタン ==========

st.markdown("")  # spacer
button_disabled = len(selected_trades) == 0

if st.button(
    "ダッシュボード生成",
    type="primary",
    use_container_width=True,
    disabled=button_disabled,
):
    with st.spinner("生成中... しばらくお待ちください"):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as out_tmp:
                output_path = out_tmp.name

            result = generate_dashboard(
                input_path=tmp_path,
                output_path=output_path,
                trade_type=selected_trades,
            )

            trade_part = "_".join(
                "輸入" if t == "Imports" else "輸出"
                for t in result["trade_types"]
            )
            st.session_state["last_result"] = result
            st.session_state["last_output_path"] = output_path
            st.session_state["last_filename"] = (
                f"{result['commodity_jp']}_{trade_part}_Dashboard.xlsx"
            )
        except Exception as e:
            st.error(
                "ダッシュボードの生成中にエラーが発生しました。\n\n"
                "**考えられる原因：**\n"
                "- 原始データの形式が想定と異なる\n"
                "- 必須列（Time / Country / Value など）が欠けている\n\n"
                "問題が続く場合は、ページをリロードしてもう一度お試しください。"
            )
            with st.expander("技術的な詳細（管理者向け）"):
                st.exception(e)

# ========== 結果表示 ==========

if "last_result" in st.session_state:
    result = st.session_state["last_result"]
    output_path = st.session_state["last_output_path"]
    filename = st.session_state["last_filename"]

    st.markdown("### 生成完了")

    st.markdown(
        f"""
        <div class="result-banner">
            <div class="result-banner-title">✓ ダッシュボードが生成されました</div>
            <p class="result-banner-desc">
                {result['summary']}<br>
                期間：{result['period']}　／　国家：{result['n_countries']}　／　
                レコード：{result['n_records']}
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_r1, col_r2 = st.columns([3, 2])
    with col_r1:
        st.markdown(
            f"<div style='color:#64748b; font-size:0.85rem; padding-top:0.5rem;'>"
            f"ファイル名： <code>{filename}</code></div>",
            unsafe_allow_html=True,
        )
    with col_r2:
        with open(output_path, "rb") as f:
            st.download_button(
                "ダウンロード（.xlsx）",
                data=f.read(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

    # 補足通知
    if result.get("new_countries"):
        st.info(
            f"**新規国家：** {', '.join(result['new_countries'])}　"
            "（英文名のまま表示されます。日文名は「国名対照」ページで登録できます）"
        )
    if result.get("provisional_months"):
        st.caption(
            f"※ 速報値月：{', '.join(result['provisional_months'])}"
        )

    # プレビュー
    with st.expander("ダッシュボードのプレビューを表示", expanded=False):
        st.caption("主要表（ランキング）のみ表示。完全版はダウンロードしてご確認ください。")

        try:
            from openpyxl import load_workbook
            wb_preview = load_workbook(output_path, data_only=True)

            for ts in result["per_trade_summaries"]:
                t = ts["trade_type"]
                sheet_name = "輸入分析ダッシュボード" if t == "Imports" else "輸出分析ダッシュボード"
                trade_label = "輸入" if t == "Imports" else "輸出"
                ws = wb_preview[sheet_name]

                t3_header_row = None
                for r in range(1, ws.max_row + 1):
                    if ws.cell(row=r, column=2).value == "順位":
                        t3_header_row = r
                        break

                if t3_header_row:
                    ranking_data = []
                    for i in range(ts["n_countries"]):
                        r = t3_header_row + 1 + i
                        ranking_data.append({
                            "順位": ws.cell(row=r, column=2).value,
                            "国名": ws.cell(row=r, column=3).value,
                            "総額(USD$1k)": ws.cell(row=r, column=4).value,
                            "総重量(TNE)": ws.cell(row=r, column=5).value,
                            "金額構成比": ws.cell(row=r, column=6).value,
                            "累計構成比": ws.cell(row=r, column=7).value,
                            "平均単価(USD/TNE)": ws.cell(row=r, column=8).value,
                            "備考": ws.cell(row=r, column=9).value or "",
                        })

                    df_ranking = pd.DataFrame(ranking_data)
                    st.markdown(f"**{trade_label}　ランキング**")
                    st.dataframe(
                        df_ranking.style.format({
                            "総額(USD$1k)": "{:,.0f}",
                            "総重量(TNE)": "{:,.0f}",
                            "金額構成比": "{:.1%}",
                            "累計構成比": "{:.1%}",
                            "平均単価(USD/TNE)": "{:.1f}",
                        }),
                        hide_index=True,
                        use_container_width=True,
                    )
        except Exception as e:
            st.caption(f"プレビュー生成失敗：{e}")

# ========== フッター ==========

footer()
